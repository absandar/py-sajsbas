from datetime import datetime
import glob
import json
import os, sys, tempfile
import sqlite3
from config import Config
from zoneinfo import ZoneInfo
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.image import Image
from openpyxl.utils import range_boundaries

from .sqlite_service import SQLiteService


class RemisionExcelBuilder:
    def __init__(self, db_path=None):
        self.sqlservice = SQLiteService()
        self.ruta_actual = os.path.dirname(os.path.abspath(__file__))
        self.image_path = os.path.join(self.ruta_actual, "images", "logo_procesa.png")
        self.image_copia_controlada = os.path.join(self.ruta_actual, "images", "copia_certificada.png")

        # === Crear libro y hoja ===
        self.wb: Workbook = Workbook()
        self.ws: Worksheet = self.wb.active  # type: ignore
        self.ws.sheet_view.zoomScale = 55
        self.ws.page_setup.scale = 40
        self.ws.page_setup.paperSize = self.ws.PAPERSIZE_LETTER
        self.ws.page_margins.left = 0.2
        self.ws.page_margins.right = 0.2
        self.ws.page_margins.top = 0.2
        self.ws.page_margins.bottom = 0.2
        self.ws.page_margins.header = 0.2
        self.ws.page_margins.footer = 0.2

        # === Estilos globales ===
        self.COLOR_GRIS = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # === Color base blanco sin bordes ===
        self.COLOR_BLANCO = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        self.sin_borde = Border()  # sin estilo alguno

        # === Aplicar fondo blanco a toda la hoja inicial ===
        # (por ejemplo, 100 columnas × 200 filas — ajustable según tus necesidades)
        for row in self.ws.iter_rows(min_row=1, max_row=200, min_col=1, max_col=100):
            for cell in row:
                cell.fill = self.COLOR_BLANCO
                cell.border = self.sin_borde

        # === Datos ===
        self.cargas_de_dia = self.sqlservice.cargas_del_dia()
        self.relaciones = self.sqlservice.relacion_sku_descripcion()
        self.retallados = self.sqlservice.retallados_del_dia()


    def _aplicar_formato_merge(self, ws: Worksheet, rango: str, texto="", fill=None, font=None, alignment=None, border=None, valor=None):
        """Aplica formato a todas las celdas de un rango mergeado"""
        ws.merge_cells(rango)
        
        # Aplicar formato a todas las celdas
        for row in ws[rango]:
            for cell in row:
                if fill:
                    cell.fill = fill
                if font:
                    cell.font = font
                if alignment:
                    cell.alignment = alignment
                if border:
                    cell.border = border
        
        # Poner el valor/texto en la primera celda
        celda_principal = ws[rango.split(":")[0]]
        if valor is not None:
            celda_principal.value = valor
        elif texto:
            celda_principal.value = texto
        
        return celda_principal
    # =============================================================
    # TABLA PRINCIPAL (REMISIÓN DE ATÚN)
    # =============================================================
    def tabla_principal(self, fila_inicial=1):
        ws = self.ws
        COLOR_GRIS = self.COLOR_GRIS
        thin_border = self.thin_border
        relaciones = self.relaciones

        # === Estructura general ===
        rangos = [
            'A1:C4', 'D1:N4', 'O1:Q4', 'A6:B6', 'D6:E6', 'F6:Q6',
            'A7:B7', 'D7:E7', 'F7:H7', 'I7:J7', 'K7:M7', 'O7:Q7'
        ]
        for rango in rangos:
            ws.merge_cells(rango)
            min_col, min_row, max_col, max_row = range_boundaries(rango)
            for row in range(min_row, max_row + 1): # type: ignore
                for col in range(min_col, max_col + 1): # type: ignore
                    cell = ws.cell(row=row, column=col)
                    left   = 'thin' if col == min_col else None
                    right  = 'thin' if col == max_col else None
                    top    = 'thin' if row == min_row else None
                    bottom = 'thin' if row == max_row else None
                    cell.border = Border(
                        left=Side(style=left),
                        right=Side(style=right),
                        top=Side(style=top),
                        bottom=Side(style=bottom)
                    )

        ws['C6'].border = thin_border
        ws['C7'].border = thin_border
        ws['N7'].border = thin_border

        # === Anchos de columna ===
        anchos = {
            'A': 8.67, 'B': 33.78, 'C': 17.56, 'D': 47.33, 'E': 22.22, 'F': 16,
            'G': 14, 'H': 13.78, 'I': 14, 'J': 12.33, 'K': 16.56, 'L': 17.44,
            'M': 21, 'N': 20, 'O': 26.67, 'P': 13.33, 'Q': 13.33
        }
        for col, width in anchos.items():
            ws.column_dimensions[col].width = width

        # === Alturas de fila ===
        alturas = {
            1: 39, 2: 39, 3: 39, 4: 39,
            5: 4.2, 6: 27, 7: 27, 8: 4.2, 9: 63
        }
        for fila, altura in alturas.items():
            ws.row_dimensions[fila].height = altura

        # === Encabezados ===
        for row in ws['A6:Q7']:
            for cell in row:
                cell.font = Font(name='Calibri', size=16, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in ws['A9:Q9']:
            for cell in row:
                cell.font = Font(name='Calibri', size=16, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.fill = COLOR_GRIS
                cell.border = thin_border

        # === Celdas grises específicas ===
        for celda in ['A6', 'D6', 'A7', 'D7', 'I7', 'N7']:
            ws[celda].fill = COLOR_GRIS

        # === Imagen ===
        if not os.path.exists(self.image_path):
            raise FileNotFoundError(f"No se encontró la imagen en: {self.image_path}")
        img = Image(self.image_path)
        ws.add_image(img, 'A1')
        img = Image(self.image_copia_controlada)

        image_width_pixels = 150   # Ancho deseado
        image_height_pixels = 150   # Alto deseado

        # Convertir tamaño a EMU
        image_width_emu = pixels_to_EMU(image_width_pixels)
        image_height_emu = pixels_to_EMU(image_height_pixels)
        size = XDRPositiveSize2D(image_width_emu, image_height_emu)

        # Definir posición (celda O1 = columna 14, fila 0 en base 0)
        target_col = 15  # Columna O (15-1)
        target_row = 0   # Fila 1 (1-1)

        # Definir offset (desplazamiento desde la esquina de la celda)
        offset_x_pixels = 35  # Offset horizontal
        offset_y_pixels = 15   # Offset vertical
        offset_x_emu = pixels_to_EMU(offset_x_pixels)
        offset_y_emu = pixels_to_EMU(offset_y_pixels)

        # Crear AnchorMarker con posición y offset
        marker = AnchorMarker(
            col=target_col, 
            colOff=offset_x_emu, 
            row=target_row, 
            rowOff=offset_y_emu
        )

        # Asignar el anchor a la imagen
        img.anchor = OneCellAnchor(_from=marker, ext=size)
        ws.add_image(img)

        # === Título y textos fijos ===
        ws['D1'] = "ALMACÉN - CÁMARAS FRIGORÍFICAS\nREMISIÓN DE ATÚN FRESCO CONGELADO"
        ws['D1'].font = Font(name='Calibri', size=26, bold=True)
        ws['D1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws['A6'] = "Folio"
        ws['A7'] = "Número de sello"
        ws['D6'] = "Cliente"
        ws['D7'] = "Placas de contenedor"
        ws['I7'] = "Factura / Pedido"
        ws['N7'] = "Fecha"

        rich_text = CellRichText()
        rich_text.append(TextBlock(InlineFont(rFont='Calibri', sz=22), "Código: "))
        rich_text.append(TextBlock(InlineFont(rFont='Calibri', sz=22, b=True), "F-AL-02\n"))
        rich_text.append(TextBlock(InlineFont(rFont='Calibri', sz=22), "Fecha: 29-09-2025\n"))
        rich_text.append(TextBlock(InlineFont(rFont='Calibri', sz=22), "Página: 1 de 1\n"))
        rich_text.append(TextBlock(InlineFont(rFont='Calibri', sz=22), "Revisión: 01\n"))
        rich_text.append(TextBlock(InlineFont(rFont='Calibri', sz=22), "Referenciado a M-AL-01"))
        ws['O1'] = rich_text
        ws['O1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        encabezados = [
            "N°", "Carga", "SKU", "Descripción", "Lote", "Tanque", "Tina",
            "Tara", "Peso Bruto", "Peso Salida", "Merma", "Peso Entrada",
            "Peso Neto Devolución", "Peso Bruto Devolución", "Observaciones",
            "MSC", "Evaluación sensorial"
        ]
        for idx, titulo in enumerate(encabezados, start=1):
            ws.cell(row=9, column=idx, value=titulo)

        # === Datos ===
        try:
            remision_general = json.loads(self.cargas_de_dia)
        except Exception as e:
            print(f"Error al decodificar cargas_del_dia: {e}")
            remision_general = {}

        if not remision_general or "cargas" not in remision_general:
            print("No hay datos de remisiones.")
            return

        # === Encabezado general ===
        ws['C6'] = remision_general.get("folio", "")
        ws['C7'] = remision_general.get("numero_sello", "")
        ws['F6'] = remision_general.get("cliente", "")
        ws['F7'] = remision_general.get("placas_contenedor", "")
        ws['K7'] = remision_general.get("factura", "")
        fecha = remision_general.get("fecha_creacion", "")
        ws['O7'] = datetime.strptime(fecha, "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y") if fecha else ""

        fila = 9
        total_peso_bruto = total_peso_salida = total_merma = total_peso_entrada = total_peso_neto_devol = 0.0
        contador = 0

        for carga in remision_general.get("cargas", []):
            primera_fila = True
            for detalle in carga.get("detalles", []):
                print("carga:", carga)
                contador += 1
                fila += 1
                ws[f"A{fila}"] = contador
                ws[f"B{fila}"] = f"CARGA {carga.get('carga', '')}: {int(carga.get('cantidad_solicitada', 0))}"
                ws[f"C{fila}"] = detalle.get("sku_talla", "")
                ws[f"D{fila}"] = relaciones.get(detalle.get("sku_talla", ""), "")
                ws[f"E{fila}"] = detalle.get("lote", "")
                ws[f"F{fila}"] = detalle.get("tanque", "")
                ws[f"G{fila}"] = detalle.get("sku_tina", "")
                ws[f"H{fila}"] = detalle.get("tara", "")
                ws[f"I{fila}"] = detalle.get("peso_bascula", "")
                ws[f"J{fila}"] = detalle.get("peso_neto", "")
                ws[f"K{fila}"] = detalle.get("merma", "")
                ws[f"L{fila}"] = detalle.get("peso_marbete", "")
                ws[f"M{fila}"] = detalle.get("peso_neto_devolucion", "")
                ws[f"N{fila}"] = detalle.get("peso_bruto_devolucion", "")
                ws[f"O{fila}"] = detalle.get("observaciones", "")
                ws[f"P{fila}"] = "SI" if detalle.get("is_msc") == 1 else ""
                ws[f"Q{fila}"] = "SI" if detalle.get("is_sensorial") == 1 else ""

                total_peso_bruto += float(detalle.get("peso_bascula") or 0)
                total_peso_salida += float(detalle.get("peso_neto") or 0)
                total_merma += float(detalle.get("merma") or 0)
                total_peso_entrada += float(detalle.get("peso_marbete") or 0)
                total_peso_neto_devol += float(detalle.get("peso_neto_devolucion") or 0)

                ws.row_dimensions[fila].height = 21.6
                for col in range(1, 18):
                    c = ws.cell(row=fila, column=col)
                    c.border = thin_border
                    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    c.font = Font(name='Calibri', size=12)
                    if primera_fila:
                        c.font = Font(name='Calibri', size=12, bold=True)
                    else:
                        c.font = Font(name='Calibri', size=12)
                primera_fila = False

        # Guardar totales en la instancia para usarlos en `totales()`
        self._total_peso_salida = total_peso_salida
        self._total_peso_neto_devol = total_peso_neto_devol
        # Guardar total de merma
        self._total_merma = total_merma
        # Guardar total peso bruto (para calcular Salida Total)
        self._total_peso_bruto = total_peso_bruto
        self._total_peso_entrada = total_peso_entrada

        # === Totales ===
        fila_total = fila + 1
        ws[f"I{fila_total}"] = total_peso_bruto
        ws[f"J{fila_total}"] = total_peso_salida
        ws[f"K{fila_total}"] = total_merma
        ws[f"L{fila_total}"] = total_peso_entrada
        ws[f"M{fila_total}"] = total_peso_neto_devol

        ws.merge_cells(f"A{fila_total}:H{fila_total}")
        ws[f"A{fila_total}"] = "Total Entregado en Remisión"
        for row in ws[f"A{fila_total}:H{fila_total}"]:
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = COLOR_GRIS
                cell.border = thin_border
                cell.font = Font(name='Calibri', size=14, bold=True)
                
        for col in range(9, 18):
            c = ws.cell(row=fila_total, column=col)
            c.font = Font(name='Calibri', size=14, bold=True)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = thin_border
            c.fill = COLOR_GRIS
            if col >= 9:
                c.number_format = '#,##0.00'
        ws.row_dimensions[fila_total].height = 24

        # Devuelve la siguiente fila libre (para construir más secciones abajo)
        return fila_total + 1


    def retallado(self, fila_inicial=11):
        ws = self.ws
        COLOR_GRIS = self.COLOR_GRIS
        thin_border = self.thin_border
        relaciones = self.relaciones
        retallados = self.retallados

        ws.row_dimensions[fila_inicial].height = 5
        fila_actual = fila_inicial + 1

        # === Título "RETALLADOS" ===
        rango_merge = f"A{fila_actual}:Q{fila_actual}"
        self._aplicar_formato_merge(
            ws, rango_merge,
            texto="RETALLADOS",
            fill=COLOR_GRIS,
            font=Font(name='Calibri', size=14, bold=True),
            alignment=Alignment(horizontal='center', vertical='center'),
            border=thin_border
        )
        fila_actual += 1

        # === Encabezados ===
        encabezados = ["N°", "SKU", "Descripción", "Lote", "Tina",
                    "Tara", "Peso Bruto", "Peso Neto", "Observaciones"]
        for idx, titulo in enumerate(encabezados, start=1):
            celda = ws.cell(row=fila_actual, column=idx, value=titulo)
            celda.font = Font(name='Calibri', size=16, bold=True)
            celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            celda.fill = COLOR_GRIS
            celda.border = thin_border
        ws.merge_cells(f"I{fila_actual}:Q{fila_actual}")

        # === Datos desde la BD ===
        fila_inicio_datos = fila_actual + 1
        if retallados:
            for i, fila_db in enumerate(retallados, start=1):
                sku_tina, sku_talla, lote, tara, peso_bascula, peso_neto, obs = fila_db
                fila_excel = fila_inicio_datos + i - 1
                datos = [
                    i,
                    sku_talla or "",
                    relaciones.get(sku_talla, ""),
                    lote or "",
                    sku_tina or "",
                    tara or 0.0,
                    peso_bascula or 0.0,
                    peso_neto or 0.0,
                    obs or ""
                ]
                for col, val in enumerate(datos, start=1):
                    c = ws.cell(row=fila_excel, column=col, value=val)
                    c.border = thin_border
                    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    c.font = Font(name='Calibri', size=12)
                ws.merge_cells(f"I{fila_excel}:Q{fila_excel}")
                ws.row_dimensions[fila_excel].height = 18
        else:
            # Si no hay datos, muestra solo una fila vacía
            fila_excel = fila_inicio_datos
            for col in range(1, 10):
                c = ws.cell(row=fila_excel, column=col, value="")
                c.border = thin_border
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.font = Font(name='Calibri', size=12)
            ws.merge_cells(f"I{fila_excel}:Q{fila_excel}")
            ws.row_dimensions[fila_excel].height = 18

        # === Fila de totales ===
        fila_total = fila_excel + 1
        ws.merge_cells(f"A{fila_total}:G{fila_total}")
        ws[f"A{fila_total}"] = "Total Peso Neto Retallado"
        ws[f"H{fila_total}"] = f"=SUM(H{fila_inicio_datos}:H{fila_excel})"
        ws[f"H{fila_total}"].number_format = '#,##0.00'
        ws.merge_cells(f"I{fila_total}:Q{fila_total}")

        for row in ws[f"A{fila_total}:Q{fila_total}"]:
            for cell in row:
                cell.font = Font(name='Calibri', size=12, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = COLOR_GRIS
                cell.border = thin_border

        return fila_total
    
    def totales(self, fila_inicial=11):
        ws = self.ws
        cargas_del_dia = json.loads(self.cargas_de_dia)
        COLOR_GRIS = self.COLOR_GRIS
        thin_border = self.thin_border

        # === función auxiliar simplificada ===
        def set_cell(rango: str, texto="", gris=False, negrita=False, merge=False, numero=False, valor=None):
            if merge:
                return self._aplicar_formato_merge(
                    ws, rango, 
                    texto=texto, 
                    fill=COLOR_GRIS if gris else None,
                    font=Font(name='Calibri', size=12, bold=negrita),
                    alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
                    border=thin_border,
                    valor=valor
                )
            else:
                # Para celdas individuales
                celda = ws[rango]
                if valor is not None:
                    celda.value = valor
                else:
                    celda.value = texto
                    
                celda.font = Font(name='Calibri', size=12, bold=negrita)
                celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                celda.border = thin_border
                if gris:
                    celda.fill = COLOR_GRIS
                if numero:
                    celda.number_format = '#,##0.00'
                return celda

        # === estructura de totales ===
        fila_actual = fila_inicial + 1

        # calcular valor neto entregado: preferir totales guardados, si no, recalcular desde cargas_del_dia
        try:
            total_salida = getattr(self, "_total_peso_salida")
            total_devol = getattr(self, "_total_peso_neto_devol")
            total_merma = getattr(self, "_total_merma")
            total_bruto = getattr(self, "_total_peso_bruto")
            total_entrada = getattr(self, "_total_peso_entrada")
            total_retallado = getattr(self, "_total_retallado")
        except Exception:
            total_salida = None
            total_devol = None
            total_merma = None
            total_bruto = None
            total_entrada = None
            total_retallado = None

        if total_salida is None or total_devol is None or total_merma is None or total_bruto is None or total_entrada is None or total_retallado is None:
            # recalcular desde cargas_del_dia
            total_salida = 0.0
            total_devol = 0.0
            total_merma = 0.0
            total_bruto = 0.0
            total_entrada = 0.0
            total_retallado = 0.0
            try:
                rem = cargas_del_dia
                for carga in rem.get("cargas", []):
                    for det in carga.get("detalles", []):
                        total_salida += float(det.get("peso_neto") or 0)
                        total_devol += float(det.get("peso_neto_devolucion") or 0)
                        total_merma += float(det.get("merma") or 0)
                        total_bruto += float(det.get("peso_bascula") or 0)
                        total_entrada += float(det.get("peso_marbete") or 0)
                # sumar pesos netos de retallados (si los tenemos en la instancia)
                try:
                    for r in (self.retallados or []):
                        # r = (sku_tina, sku_talla, lote, tara, peso_bascula, peso_neto, obs)
                        total_retallado += float((r[5] or 0))
                except Exception:
                    total_retallado = total_retallado or 0.0
            except Exception:
                total_salida = 0.0
                total_devol = 0.0
                total_merma = 0.0
                total_bruto = 0.0
                total_entrada = 0.0
                total_retallado = 0.0

        # fórmula solicitada:
        # total_peso_neto_entregado = total_peso_salida - total_peso_neto_devolucion - total_peso_neto_retallado
        total_peso_neto_entregado = total_salida - total_devol - total_retallado
        # Salida Total = total de peso de entrada
        salida_total = total_entrada

        set_cell(f"A{fila_actual}:C{fila_actual}", "Total Peso Neto Entregado", gris=True, negrita=True, merge=True)
        set_cell(f"D{fila_actual}", valor=total_peso_neto_entregado, numero=True)
        set_cell(f"H{fila_actual}:I{fila_actual}", "Merma", gris=True, negrita=True, merge=True)
        set_cell(f"J{fila_actual}:K{fila_actual}", valor=total_merma, numero=True, merge=True)
        set_cell(f"O{fila_actual}", "Salida Total", gris=True, negrita=True)
        set_cell(f"P{fila_actual}:Q{fila_actual}", valor=salida_total, numero=True, merge=True)

        ws.row_dimensions[fila_actual].height = 18
        fila_actual += 1
        
        # === Bloque de 2 filas unidas (debajo de la fila de totales) ===
        fila_bloque_ini = fila_actual + 1
        fila_bloque_fin = fila_actual + 2

        # Altura de ambas filas
        ws.row_dimensions[fila_bloque_ini].height = 47.25
        ws.row_dimensions[fila_bloque_fin].height = 47.25

        # Usar la nueva función para los merges
        self._aplicar_formato_merge(
            ws, f"A{fila_bloque_ini}:C{fila_bloque_fin}",
            texto="Observaciones",
            fill=COLOR_GRIS,
            font=Font(name='Calibri', size=12, bold=True),
            alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
            border=thin_border
        )

        self._aplicar_formato_merge(
            ws, f"D{fila_bloque_ini}:Q{fila_bloque_fin}",
            font=Font(name='Calibri', size=16, bold=True),
            alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
            border=thin_border
        )
        ws[f"D{fila_bloque_ini}"] = cargas_del_dia['observaciones_cabecera']
        # Actualiza y devuelve la última fila usada
        fila_actual = fila_bloque_fin
        fila_vacia_1 = fila_actual + 1
        fila_vacia_2 = fila_actual + 2

        ws.row_dimensions[fila_vacia_1].height = 31.5
        ws.row_dimensions[fila_vacia_2].height = 31.5

        # (sin contenido, solo altura)
        fila_actual = fila_vacia_2

        # === fila donde firman ===
        fila_bloque_ini = fila_actual + 1
        fila_bloque_fin = fila_actual + 2

        ws.row_dimensions[fila_bloque_ini].height = 31.5
        ws.row_dimensions[fila_bloque_fin].height = 31.5

        # Usar la nueva función para las firmas
        self._aplicar_formato_merge(
            ws, f"A{fila_bloque_ini}:E{fila_bloque_fin}",
            texto="Nombre y Firma\nRecibió",
            fill=COLOR_GRIS,
            font=Font(name='Calibri', size=14, bold=True),
            alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
            border=thin_border
        )

        self._aplicar_formato_merge(
            ws, f"F{fila_bloque_ini}:K{fila_bloque_fin}",
            texto="Nombre y Firma\nEntregó",
            fill=COLOR_GRIS,
            font=Font(name='Calibri', size=14, bold=True),
            alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
            border=thin_border
        )

        self._aplicar_formato_merge(
            ws, f"L{fila_bloque_ini}:Q{fila_bloque_fin}",
            texto="Nombre y Firma\nAutorizó",
            fill=COLOR_GRIS,
            font=Font(name='Calibri', size=14, bold=True),
            alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
            border=thin_border
        )

        # Actualiza última fila
        fila_actual = fila_bloque_fin
        return fila_actual
    # =============================================================
    # Guardar archivo final
    # =============================================================
    def guardar(self, nombre="ejemplo.xlsx"):
        # === 1. Intentar leer ruta desde configuraciones_varias.json ===
        carpeta_destino = None
        try:
            with open(Config.LOCAL_CONFIGS, 'r') as f:
                data = json.load(f)
                carpeta_destino = data.get("EXPORT_FOLDER")
        except:
            carpeta_destino = None

        # === 2. Si no existe en JSON, usar variable de entorno ===
        if not carpeta_destino:
            carpeta_destino = Config.EXPORT_FOLDER

        # === 3. Si tampoco está en variable de entorno, usar Descargas ===
        if not carpeta_destino:
            carpeta_destino = os.path.join(os.path.expanduser("~"), "Downloads")

        # === Crear carpeta si no existe ===
        os.makedirs(carpeta_destino, exist_ok=True)

        # === Preparar limpieza de archivos previos ===
        prefijo = nombre[:21]
        patron_busqueda = os.path.join(carpeta_destino, f"{prefijo}*")
        archivos_existentes = glob.glob(patron_busqueda)

        for archivo in archivos_existentes:
            try:
                os.remove(archivo)
            except:
                pass

        # === Guardar archivo ===
        ruta_completa = os.path.join(carpeta_destino, nombre)
        self.wb.save(ruta_completa)
        return ruta_completa